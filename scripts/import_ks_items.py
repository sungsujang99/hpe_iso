#!/usr/bin/env python
"""
KS 인증 사내문서 관리대장 엑셀 데이터 Import 스크립트
"""
import os
import sys
import django

# Django 설정
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.development')
django.setup()

from apps.ks_certification.models import KSCertificationItem, KSCertificationHistory
from apps.accounts.models import User
import openpyxl
from datetime import datetime


def import_ks_items_from_excel(excel_file='KS 인증 사내문서 관리대장_251226.xlsx'):
    """엑셀 파일에서 KS 인증 품목 데이터 import"""
    
    print(f"📊 KS 인증 품목 데이터 Import 시작...")
    print(f"파일: {excel_file}\n")
    
    # 관리자 계정 가져오기 (생성자로 사용)
    try:
        admin_user = User.objects.filter(role=User.Role.ADMIN).first()
        if not admin_user:
            print("❌ 관리자 계정을 찾을 수 없습니다.")
            return
    except Exception as e:
        print(f"❌ 관리자 계정 조회 실패: {e}")
        return
    
    # 엑셀 파일 읽기
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Sheet1']
    except FileNotFoundError:
        print(f"❌ 파일을 찾을 수 없습니다: {excel_file}")
        return
    except Exception as e:
        print(f"❌ 엑셀 파일 읽기 실패: {e}")
        return
    
    # 데이터 추출
    items = []
    current_item = None
    
    for row_idx in range(8, ws.max_row + 1):
        no = ws.cell(row=row_idx, column=2).value
        barcode = ws.cell(row=row_idx, column=3).value
        name = ws.cell(row=row_idx, column=6).value
        
        # NO와 관리번호가 있으면 새 항목 시작
        if no is not None and barcode is not None:
            # 이전 항목 저장
            if current_item:
                items.append(current_item)
            
            # 새 항목 생성
            current_item = {
                'no': str(no).strip(),
                'barcode': str(barcode).strip(),
                'name': str(name).strip() if name else ''
            }
        # NO와 관리번호가 없지만 이름이 있으면 이전 항목에 추가 (여러 행에 걸친 이름)
        elif current_item and name is not None:
            current_item['name'] += ' ' + str(name).strip()
    
    # 마지막 항목 저장
    if current_item:
        items.append(current_item)
    
    wb.close()
    
    print(f"✅ 엑셀에서 {len(items)}개 품목 추출 완료\n")
    
    # DB에 저장
    created_count = 0
    updated_count = 0
    error_count = 0
    
    for idx, item_data in enumerate(items, 1):
        barcode = item_data['barcode']
        name = item_data['name']
        
        try:
            # 기존 품목 확인
            existing_item = KSCertificationItem.objects.filter(barcode=barcode).first()
            
            if existing_item:
                # 업데이트
                existing_item.name = name
                existing_item.updated_by = admin_user
                existing_item.save()
                
                # 이력 기록
                KSCertificationHistory.objects.create(
                    item=existing_item,
                    action_type=KSCertificationHistory.ActionType.UPDATE,
                    action_description=f'엑셀 데이터로 업데이트',
                    new_value={'name': name},
                    created_by=admin_user
                )
                
                updated_count += 1
                print(f"[{idx:3d}] 업데이트: {barcode} - {name[:50]}")
            else:
                # 새로 생성
                new_item = KSCertificationItem.objects.create(
                    barcode=barcode,
                    name=name,
                    status=KSCertificationItem.Status.ACTIVE,
                    quantity=0,
                    unit='EA',
                    created_by=admin_user,
                    updated_by=admin_user
                )
                
                # 이력 기록
                KSCertificationHistory.objects.create(
                    item=new_item,
                    action_type=KSCertificationHistory.ActionType.REGISTER,
                    action_description=f'엑셀 데이터로 신규 등록',
                    new_value={'barcode': barcode, 'name': name},
                    created_by=admin_user
                )
                
                created_count += 1
                print(f"[{idx:3d}] 신규: {barcode} - {name[:50]}")
        
        except Exception as e:
            error_count += 1
            print(f"[{idx:3d}] 오류: {barcode} - {e}")
    
    print(f"\n" + "="*80)
    print(f"✅ Import 완료!")
    print(f"  - 신규 생성: {created_count}개")
    print(f"  - 업데이트: {updated_count}개")
    print(f"  - 오류: {error_count}개")
    print(f"  - 총: {len(items)}개")
    print(f"="*80)


if __name__ == '__main__':
    import_ks_items_from_excel()
