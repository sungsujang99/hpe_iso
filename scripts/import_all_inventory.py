#!/usr/bin/env python
"""
통합 재고 데이터 Import 스크립트
- HP-KSTC: KS 인증 사내문서 관리대장 (110개)
- HP-P10: 계측장비 재고조사 관리대장 (93개)
- HP-PRT: 사내부품 재고관리 (5개)
- HP-SUP: 사내소모품 재고관리 (52개)
"""
import os
import sys
import django

# Django 설정
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.development')
django.setup()

from apps.inventory.models import InventoryItem, StockTransaction
from apps.accounts.models import User
import openpyxl


def import_ks_certification():
    """KS 인증 사내문서 관리대장 (110개)"""
    print("\n📋 1. KS 인증 사내문서 관리대장 Import 시작...")
    print("="*80)
    
    wb = openpyxl.load_workbook('KS 인증 사내문서 관리대장_251226.xlsx')
    ws = wb['Sheet1']
    
    items = []
    current_item = None
    
    for row_idx in range(8, ws.max_row + 1):
        no = ws.cell(row=row_idx, column=2).value
        barcode = ws.cell(row=row_idx, column=3).value
        name = ws.cell(row=row_idx, column=6).value
        
        if no is not None and barcode is not None:
            if current_item:
                items.append(current_item)
            
            current_item = {
                'barcode': str(barcode).strip(),
                'name': str(name).strip() if name else ''
            }
        elif current_item and name is not None:
            current_item['name'] += ' ' + str(name).strip()
    
    if current_item:
        items.append(current_item)
    
    wb.close()
    
    created_count = 0
    for item_data in items:
        barcode = item_data['barcode']
        name = item_data['name']
        
        if not InventoryItem.objects.filter(barcode=barcode).exists():
            InventoryItem.objects.create(
                barcode=barcode,
                item_code=barcode,
                name=name,
                # item_type은 save()에서 자동 설정됨 (HP-KSTC -> KS_CERTIFICATION)
                unit='EA',
                current_quantity=0,
                certification_body='한국표준협회'
            )
            created_count += 1
    
    print(f"✅ KS 인증: {created_count}개 import 완료")
    return created_count


def import_measurement_equipment():
    """계측장비 재고조사 관리대장 (93개)"""
    print("\n🔧 2. 계측장비 재고조사 관리대장 Import 시작...")
    print("="*80)
    
    wb = openpyxl.load_workbook('계측장비 재고조사 관리대장_251224.xlsx')
    ws = wb['계측장비관리대장']
    
    items = []
    
    for row_idx in range(1, ws.max_row + 1):
        no = ws.cell(row=row_idx, column=1).value
        barcode = ws.cell(row=row_idx, column=2).value
        name = ws.cell(row=row_idx, column=5).value
        
        if no and barcode and isinstance(barcode, str) and barcode.startswith('HP-P'):
            items.append({
                'barcode': str(barcode).strip(),
                'name': str(name).strip() if name else ''
            })
    
    wb.close()
    
    created_count = 0
    for item_data in items:
        barcode = item_data['barcode']
        name = item_data['name']
        
        # 사양 파싱 (예: "PRESSURE GAUGE 1000kgf/cm2")
        specifications = name
        measurement_range = ''
        
        if not InventoryItem.objects.filter(barcode=barcode).exists():
            InventoryItem.objects.create(
                barcode=barcode,
                item_code=barcode,
                name=name,
                # item_type은 save()에서 자동 설정됨 (HP-P10 -> MEASUREMENT)
                unit='EA',
                current_quantity=0,
                equipment_type='pressure',  # 압력계
                specification=specifications,
                measurement_range=measurement_range,
                calibration_required=True,
                calibration_cycle_months=12
            )
            created_count += 1
    
    print(f"✅ 계측장비: {created_count}개 import 완료")
    return created_count


def import_parts():
    """사내부품 재고관리 (5개)"""
    print("\n🔩 3. 사내부품 재고관리 Import 시작...")
    print("="*80)
    
    wb = openpyxl.load_workbook('재고관리 리스트 (PRT).xlsx')
    ws = wb['사내부품(PRT)-001']
    
    items = []
    
    for row_idx in range(1, ws.max_row + 1):
        barcode = ws.cell(row=row_idx, column=4).value
        name = ws.cell(row=row_idx, column=6).value
        received = ws.cell(row=row_idx, column=8).value
        issued = ws.cell(row=row_idx, column=9).value
        
        if barcode and isinstance(barcode, str) and barcode.startswith('HP-PRT'):
            items.append({
                'barcode': str(barcode).strip(),
                'name': str(name).strip() if name else '',
                'received': received if isinstance(received, (int, float)) else 0,
                'issued': issued if isinstance(issued, (int, float)) else 0
            })
    
    wb.close()
    
    created_count = 0
    for item_data in items:
        barcode = item_data['barcode']
        name = item_data['name']
        received = item_data['received']
        issued = item_data['issued']
        current = received - issued
        
        if not InventoryItem.objects.filter(barcode=barcode).exists():
            InventoryItem.objects.create(
                barcode=barcode,
                item_code=barcode,
                name=name,
                # item_type은 save()에서 자동 설정됨 (HP-PRT -> PARTS)
                unit='EA',
                current_quantity=current,
                received_quantity=received,
                issued_quantity=issued
            )
            created_count += 1
    
    print(f"✅ 사내부품: {created_count}개 import 완료")
    return created_count


def import_supplies():
    """사내소모품 재고관리 (52개)"""
    print("\n📦 4. 사내소모품 재고관리 Import 시작...")
    print("="*80)
    
    wb = openpyxl.load_workbook('재고관리 리스트 (SUP).xlsx')
    ws = wb['사내부품(PRT)-001']
    
    items = []
    
    for row_idx in range(1, ws.max_row + 1):
        barcode = ws.cell(row=row_idx, column=4).value
        name = ws.cell(row=row_idx, column=6).value
        received = ws.cell(row=row_idx, column=8).value
        issued = ws.cell(row=row_idx, column=9).value
        
        if barcode and isinstance(barcode, str) and barcode.startswith('HP-SUP'):
            items.append({
                'barcode': str(barcode).strip(),
                'name': str(name).strip() if name else '',
                'received': received if isinstance(received, (int, float)) else 0,
                'issued': issued if isinstance(issued, (int, float)) else 0
            })
    
    wb.close()
    
    created_count = 0
    for item_data in items:
        barcode = item_data['barcode']
        name = item_data['name']
        received = item_data['received']
        issued = item_data['issued']
        current = received - issued
        
        if not InventoryItem.objects.filter(barcode=barcode).exists():
            InventoryItem.objects.create(
                barcode=barcode,
                item_code=barcode,
                name=name,
                # item_type은 save()에서 자동 설정됨 (HP-SUP -> SUPPLIES)
                unit='EA',
                current_quantity=current,
                received_quantity=received,
                issued_quantity=issued
            )
            created_count += 1
    
    print(f"✅ 사내소모품: {created_count}개 import 완료")
    return created_count


if __name__ == '__main__':
    print("\n" + "="*80)
    print("🚀 HPE 통합 재고관리 시스템 - 전체 데이터 Import")
    print("="*80)
    
    total_count = 0
    
    # 1. KS 인증 사내문서 관리대장
    total_count += import_ks_certification()
    
    # 2. 계측장비 재고조사 관리대장
    total_count += import_measurement_equipment()
    
    # 3. 사내부품 재고관리
    total_count += import_parts()
    
    # 4. 사내소모품 재고관리
    total_count += import_supplies()
    
    print("\n" + "="*80)
    print(f"🎉 전체 Import 완료!")
    print(f"총 {total_count}개 품목이 등록되었습니다.")
    print("="*80)
    
    # 통계 출력
    print("\n📊 품목 유형별 통계:")
    from django.db.models import Count
    stats = InventoryItem.objects.values('item_type').annotate(count=Count('id'))
    for stat in stats:
        item_type = dict(InventoryItem.ItemType.choices).get(stat['item_type'], stat['item_type'])
        print(f"  - {item_type}: {stat['count']}개")
    
    print(f"\n전체 품목 수: {InventoryItem.objects.count()}개")
