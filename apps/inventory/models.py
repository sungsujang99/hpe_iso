"""
Inventory Management Models
바코드 기반 재고관리 시스템
"""
from django.db import models
from django.conf import settings
from django.utils.translation import gettext_lazy as _
from django.core.validators import MinValueValidator
import uuid
import openpyxl
from pathlib import Path


class Warehouse(models.Model):
    """창고"""
    
    code = models.CharField(_('창고 코드'), max_length=20, unique=True)
    name = models.CharField(_('창고명'), max_length=100)
    address = models.TextField(_('주소'), blank=True)
    manager = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='managed_warehouses',
        verbose_name=_('담당자')
    )
    is_active = models.BooleanField(_('활성화'), default=True)
    created_at = models.DateTimeField(auto_now_add=True)
    
    class Meta:
        verbose_name = _('창고')
        verbose_name_plural = _('창고')
        ordering = ['code']
    
    def __str__(self):
        return f"{self.code} - {self.name}"


class Location(models.Model):
    """창고 내 위치 (구역)"""
    
    warehouse = models.ForeignKey(
        Warehouse,
        on_delete=models.CASCADE,
        related_name='locations',
        verbose_name=_('창고')
    )
    code = models.CharField(_('위치 코드'), max_length=30)
    name = models.CharField(_('위치명'), max_length=100)
    description = models.TextField(_('설명'), blank=True)
    barcode = models.CharField(_('바코드'), max_length=50, unique=True, blank=True, null=True)
    is_active = models.BooleanField(_('활성화'), default=True)
    
    class Meta:
        verbose_name = _('위치')
        verbose_name_plural = _('위치')
        ordering = ['warehouse', 'code']
        unique_together = ['warehouse', 'code']
    
    def __str__(self):
        return f"{self.warehouse.code}-{self.code}"
    
    def save(self, *args, **kwargs):
        if not self.barcode:
            self.barcode = f"LOC-{self.warehouse.code}-{self.code}"
        super().save(*args, **kwargs)


class ItemCategory(models.Model):
    """품목 카테고리"""
    
    code = models.CharField(_('카테고리 코드'), max_length=20, unique=True)
    name = models.CharField(_('카테고리명'), max_length=100)
    parent = models.ForeignKey(
        'self',
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='children',
        verbose_name=_('상위 카테고리')
    )
    description = models.TextField(_('설명'), blank=True)
    is_active = models.BooleanField(_('활성화'), default=True)
    
    class Meta:
        verbose_name = _('품목 카테고리')
        verbose_name_plural = _('품목 카테고리')
        ordering = ['code']
    
    def __str__(self):
        return f"{self.code} - {self.name}"


class InventoryItem(models.Model):
    """재고 품목"""
    
    class ItemType(models.TextChoices):
        MATERIAL = 'material', _('자재')
        PRODUCT = 'product', _('제품')
        CONSUMABLE = 'consumable', _('소모품')
        EQUIPMENT = 'equipment', _('장비')
        KS_CERTIFICATION = 'ks_certification', _('KS인증')
        MEASUREMENT = 'measurement', _('계측장비')
        PARTS = 'parts', _('사내부품')
        SUPPLIES = 'supplies', _('사내소모품')
    
    # Identification
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)
    item_code = models.CharField(_('품목 코드'), max_length=50, unique=True)
    barcode = models.CharField(_('바코드'), max_length=100, unique=True, blank=True, null=True)
    
    # Basic Info
    name = models.CharField(_('품목명'), max_length=200)
    description = models.TextField(_('설명'), blank=True)
    category = models.ForeignKey(
        ItemCategory,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='items',
        verbose_name=_('카테고리')
    )
    item_type = models.CharField(
        _('품목 유형'),
        max_length=20,
        choices=ItemType.choices,
        default=ItemType.MATERIAL,
    )
    
    # Specifications
    unit = models.CharField(_('단위'), max_length=20, default='EA')  # EA, KG, M, L, etc.
    specification = models.CharField(_('규격'), max_length=200, blank=True)
    manufacturer = models.CharField(_('제조사'), max_length=100, blank=True)
    supplier = models.CharField(_('공급업체'), max_length=100, blank=True)
    
    # Serial & Inspection
    serial_number = models.CharField(_('시리얼번호'), max_length=100, blank=True)
    inspection_required = models.BooleanField(_('점검여부'), default=False)
    inspection_due_date = models.DateField(_('점검예정일자'), null=True, blank=True)
    
    # KS Certification (HP-KSTC only)
    ks_standard_number = models.CharField(_('KS규격번호'), max_length=50, blank=True)
    certification_date = models.DateField(_('인증일자'), null=True, blank=True)
    expiry_date = models.DateField(_('만료일자'), null=True, blank=True)
    certification_body = models.CharField(_('인증기관'), max_length=100, blank=True)
    certification_status = models.CharField(_('인증상태'), max_length=20, blank=True)  # active, expired, suspended
    
    # Measurement Equipment (HP-P10/P20 only)
    equipment_type = models.CharField(_('장비유형'), max_length=50, blank=True)  # pressure, temperature, etc.
    model_number = models.CharField(_('모델번호'), max_length=100, blank=True)
    measurement_range = models.CharField(_('측정범위'), max_length=100, blank=True)
    accuracy = models.CharField(_('정확도'), max_length=50, blank=True)
    calibration_required = models.BooleanField(_('검정필요'), default=False)
    calibration_cycle_months = models.IntegerField(_('검정주기(개월)'), null=True, blank=True)
    last_calibration_date = models.DateField(_('최근검정일'), null=True, blank=True)
    next_calibration_date = models.DateField(_('다음검정일'), null=True, blank=True)
    calibration_agency = models.CharField(_('검정기관'), max_length=100, blank=True)
    equipment_status = models.CharField(_('장비상태'), max_length=20, blank=True)  # in_use, maintenance, calibration, broken
    responsible_person = models.CharField(_('담당자'), max_length=50, blank=True)
    department = models.CharField(_('사용부서'), max_length=100, blank=True)
    
    # Purchase Info (all types)
    purchase_date = models.DateField(_('구매일'), null=True, blank=True)
    purchase_price = models.DecimalField(_('구매가격'), max_digits=12, decimal_places=2, null=True, blank=True)
    warranty_expiry = models.DateField(_('보증만료일'), null=True, blank=True)
    
    # Stock Movement (PRT/SUP)
    received_quantity = models.DecimalField(_('입고수량'), max_digits=12, decimal_places=2, default=0)
    issued_quantity = models.DecimalField(_('출고수량'), max_digits=12, decimal_places=2, default=0)
    
    # Inventory
    current_quantity = models.DecimalField(
        _('현재 수량'),
        max_digits=12,
        decimal_places=2,
        default=0,
        validators=[MinValueValidator(0)]
    )
    safety_stock = models.DecimalField(
        _('안전재고'),
        max_digits=12,
        decimal_places=2,
        default=0,
        validators=[MinValueValidator(0)]
    )
    
    # Location
    default_location = models.ForeignKey(
        Location,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='items',
        verbose_name=_('기본 위치')
    )
    
    # Pricing
    unit_price = models.DecimalField(
        _('단가'),
        max_digits=12,
        decimal_places=2,
        default=0,
        validators=[MinValueValidator(0)]
    )
    
    # Image
    image = models.ImageField(_('이미지'), upload_to='inventory/items/', blank=True, null=True)
    
    # Status
    is_active = models.BooleanField(_('활성화'), default=True)
    
    # ISO Connection (HP-QP-710)
    iso_document = models.CharField(_('ISO 문서번호'), max_length=50, blank=True)
    
    # Timestamps
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)
    created_by = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.SET_NULL,
        null=True,
        related_name='created_items',
        verbose_name=_('등록자')
    )
    
    class Meta:
        verbose_name = _('재고 품목')
        verbose_name_plural = _('재고 품목')
        ordering = ['item_code']
        indexes = [
            models.Index(fields=['barcode']),
            models.Index(fields=['item_code']),
            models.Index(fields=['name']),
        ]
    
    def __str__(self):
        return f"{self.item_code} - {self.name}"
    
    def save(self, *args, **kwargs):
        # Barcode가 없으면 item_code로 자동 생성
        if not self.barcode:
            self.barcode = f"ITM-{self.item_code}"
        
        # Barcode 패턴에 따라 ItemType 자동 설정
        if self.barcode:
            if self.barcode.startswith('HP-KSTC-'):
                self.item_type = self.ItemType.KS_CERTIFICATION
            elif self.barcode.startswith('HP-P10-') or self.barcode.startswith('HP-P20-'):
                self.item_type = self.ItemType.MEASUREMENT
            elif self.barcode.startswith('HP-PRT-'):
                self.item_type = self.ItemType.PARTS
            elif self.barcode.startswith('HP-SUP-'):
                self.item_type = self.ItemType.SUPPLIES
        
        super().save(*args, **kwargs)
    
    @property
    def is_low_stock(self):
        """안전재고 미달 여부"""
        return self.current_quantity <= self.safety_stock
    
    @property
    def stock_status(self):
        """재고 상태"""
        if self.current_quantity <= 0:
            return 'out_of_stock'
        elif self.is_low_stock:
            return 'low_stock'
        return 'in_stock'
    
    @property
    def is_expired(self):
        """인증 만료 여부 (KS Certification only)"""
        from django.utils import timezone
        if self.expiry_date:
            return self.expiry_date < timezone.now().date()
        return False
    
    @property
    def is_calibration_due(self):
        """검정 기한 도래 여부 (Measurement Equipment only)"""
        from django.utils import timezone
        if self.next_calibration_date:
            return self.next_calibration_date <= timezone.now().date()
        return False
    
    @property
    def is_calibration_overdue(self):
        """검정 기한 초과 여부 (Measurement Equipment only)"""
        from django.utils import timezone
        if self.next_calibration_date:
            return self.next_calibration_date < timezone.now().date()
        return False


class StockTransaction(models.Model):
    """재고 입/출고 거래"""
    
    class TransactionType(models.TextChoices):
        IN = 'in', _('입고')
        OUT = 'out', _('출고')
        ADJUST = 'adjust', _('조정')
        TRANSFER = 'transfer', _('이동')
        RETURN = 'return', _('반품')
    
    # Identification
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)
    transaction_number = models.CharField(_('거래번호'), max_length=50, unique=True)
    
    # Transaction Info
    item = models.ForeignKey(
        InventoryItem,
        on_delete=models.PROTECT,
        related_name='transactions',
        verbose_name=_('품목')
    )
    transaction_type = models.CharField(
        _('거래 유형'),
        max_length=20,
        choices=TransactionType.choices,
    )
    
    # Quantity
    quantity = models.DecimalField(
        _('수량'),
        max_digits=12,
        decimal_places=2,
        validators=[MinValueValidator(0.01)]
    )
    before_quantity = models.DecimalField(
        _('거래 전 수량'),
        max_digits=12,
        decimal_places=2,
    )
    after_quantity = models.DecimalField(
        _('거래 후 수량'),
        max_digits=12,
        decimal_places=2,
    )
    
    # Location
    location = models.ForeignKey(
        Location,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='transactions',
        verbose_name=_('위치')
    )
    to_location = models.ForeignKey(
        Location,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='transfer_transactions',
        verbose_name=_('이동 위치')
    )
    
    # Reference
    reference_number = models.CharField(_('참조번호'), max_length=100, blank=True)  # PO번호, 출고요청번호 등
    remarks = models.TextField(_('비고'), blank=True)
    
    # User & Time
    performed_by = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.PROTECT,
        related_name='stock_transactions',
        verbose_name=_('담당자')
    )
    created_at = models.DateTimeField(_('거래일시'), auto_now_add=True)
    
    # Scan Info
    scanned_barcode = models.CharField(_('스캔된 바코드'), max_length=100, blank=True)
    scan_device = models.CharField(_('스캔 장치'), max_length=100, blank=True)
    
    class Meta:
        verbose_name = _('재고 거래')
        verbose_name_plural = _('재고 거래')
        ordering = ['-created_at']
        indexes = [
            models.Index(fields=['transaction_number']),
            models.Index(fields=['item', 'created_at']),
            models.Index(fields=['transaction_type', 'created_at']),
        ]
    
    def __str__(self):
        return f"{self.transaction_number} - {self.item.name} ({self.transaction_type})"
    
    def save(self, *args, **kwargs):
        if not self.transaction_number:
            from django.utils import timezone
            timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
            self.transaction_number = f"TRX-{timestamp}-{uuid.uuid4().hex[:6].upper()}"
        super().save(*args, **kwargs)


class StockAlert(models.Model):
    """재고 알림"""
    
    class AlertType(models.TextChoices):
        LOW_STOCK = 'low_stock', _('안전재고 미달')
        OUT_OF_STOCK = 'out_of_stock', _('재고 소진')
        OVERSTOCK = 'overstock', _('과잉 재고')
    
    item = models.ForeignKey(
        InventoryItem,
        on_delete=models.CASCADE,
        related_name='alerts',
        verbose_name=_('품목')
    )
    alert_type = models.CharField(
        _('알림 유형'),
        max_length=20,
        choices=AlertType.choices,
    )
    message = models.TextField(_('알림 메시지'))
    current_quantity = models.DecimalField(_('현재 수량'), max_digits=12, decimal_places=2)
    threshold_quantity = models.DecimalField(_('기준 수량'), max_digits=12, decimal_places=2)
    
    is_resolved = models.BooleanField(_('해결됨'), default=False)
    resolved_at = models.DateTimeField(_('해결일시'), null=True, blank=True)
    resolved_by = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='resolved_alerts',
        verbose_name=_('해결자')
    )
    
    created_at = models.DateTimeField(auto_now_add=True)
    
    class Meta:
        verbose_name = _('재고 알림')
        verbose_name_plural = _('재고 알림')
        ordering = ['-created_at']
    
    def __str__(self):
        return f"{self.item.name} - {self.get_alert_type_display()}"


class InventoryCount(models.Model):
    """재고 실사"""
    
    class Status(models.TextChoices):
        DRAFT = 'draft', _('작성중')
        IN_PROGRESS = 'in_progress', _('진행중')
        COMPLETED = 'completed', _('완료')
        CANCELLED = 'cancelled', _('취소')
    
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)
    count_number = models.CharField(_('실사번호'), max_length=50, unique=True)
    
    warehouse = models.ForeignKey(
        Warehouse,
        on_delete=models.PROTECT,
        related_name='inventory_counts',
        verbose_name=_('창고')
    )
    
    status = models.CharField(
        _('상태'),
        max_length=20,
        choices=Status.choices,
        default=Status.DRAFT,
    )
    
    count_date = models.DateField(_('실사일'))
    remarks = models.TextField(_('비고'), blank=True)
    
    created_by = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.PROTECT,
        related_name='created_counts',
        verbose_name=_('생성자')
    )
    approved_by = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name='approved_counts',
        verbose_name=_('승인자')
    )
    
    created_at = models.DateTimeField(auto_now_add=True)
    completed_at = models.DateTimeField(_('완료일시'), null=True, blank=True)
    
    class Meta:
        verbose_name = _('재고 실사')
        verbose_name_plural = _('재고 실사')
        ordering = ['-created_at']
    
    def __str__(self):
        return f"{self.count_number} - {self.warehouse.name}"


class InventoryCountItem(models.Model):
    """재고 실사 품목"""
    
    inventory_count = models.ForeignKey(
        InventoryCount,
        on_delete=models.CASCADE,
        related_name='items',
        verbose_name=_('실사')
    )
    item = models.ForeignKey(
        InventoryItem,
        on_delete=models.PROTECT,
        verbose_name=_('품목')
    )
    
    system_quantity = models.DecimalField(_('시스템 수량'), max_digits=12, decimal_places=2)
    counted_quantity = models.DecimalField(
        _('실사 수량'),
        max_digits=12,
        decimal_places=2,
        null=True,
        blank=True
    )
    difference = models.DecimalField(
        _('차이'),
        max_digits=12,
        decimal_places=2,
        null=True,
        blank=True
    )
    remarks = models.TextField(_('비고'), blank=True)
    
    counted_by = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        verbose_name=_('실사자')
    )
    counted_at = models.DateTimeField(_('실사일시'), null=True, blank=True)
    
    class Meta:
        verbose_name = _('실사 품목')
        verbose_name_plural = _('실사 품목')
        unique_together = ['inventory_count', 'item']
    
    def save(self, *args, **kwargs):
        if self.counted_quantity is not None:
            self.difference = self.counted_quantity - self.system_quantity
        super().save(*args, **kwargs)


# ==============================================================================
# 엑셀 기반 문서 관리 모델
# ==============================================================================

class ExcelMasterDocument(models.Model):
    """엑셀 마스터 문서 (각 엑셀 파일)"""
    
    class DocType(models.TextChoices):
        KS_CERT = 'ks_cert', 'KS 인증 사내문서 관리대장'
        MEASUREMENT = 'measurement', '계측장비 재고조사 관리대장'
        PARTS = 'parts', '재고관리 리스트 (PRT)'
        SUPPLIES = 'supplies', '재고관리 리스트 (SUP)'
    
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)
    doc_type = models.CharField(_('문서유형'), max_length=20, choices=DocType.choices, unique=True)
    title = models.CharField(_('문서제목'), max_length=200)
    file_path = models.CharField(_('파일경로'), max_length=500)
    sheet_name = models.CharField(_('시트명'), max_length=100, default='Sheet1')
    
    # 엑셀 구조 정보
    header_row = models.IntegerField(_('헤더행'), default=6)
    data_start_row = models.IntegerField(_('데이터시작행'), default=8)
    barcode_column = models.IntegerField(_('바코드컬럼'), default=2)  # B열
    name_column = models.IntegerField(_('이름컬럼'), default=5)  # E열
    
    # 추가 컬럼 (문서 유형별로 다름)
    extra_columns = models.JSONField(_('추가컬럼정보'), default=dict, blank=True)
    # 예: {'received': 8, 'issued': 9, 'current': 10}
    
    total_items = models.IntegerField(_('총항목수'), default=0)
    last_updated = models.DateTimeField(_('최종수정일시'), auto_now=True)
    created_at = models.DateTimeField(_('등록일시'), auto_now_add=True)
    
    class Meta:
        db_table = 'excel_master_documents'
        verbose_name = _('엑셀 마스터 문서')
        verbose_name_plural = _('엑셀 마스터 문서')
    
    def __str__(self):
        return self.title
    
    def get_file_path(self):
        """실제 파일 경로 반환"""
        from pathlib import Path
        return Path(settings.MEDIA_ROOT) / self.file_path
    
    def find_item_row(self, barcode):
        """바코드로 엑셀 행 찾기"""
        import openpyxl
        
        file_path = self.get_file_path()
        if not file_path.exists():
            return None
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb[self.sheet_name]
        
        for row_idx in range(self.data_start_row, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=self.barcode_column).value
            if cell_value and str(cell_value).strip() == barcode:
                wb.close()
                return row_idx
        
        wb.close()
        return None
    
    def update_item(self, barcode, updates):
        """
        바코드로 항목 찾아서 업데이트
        updates: {column_key: value} 형태
        예: {'received': 30, 'issued': 5, 'current': 25}
        """
        import openpyxl
        
        file_path = self.get_file_path()
        if not file_path.exists():
            return False
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb[self.sheet_name]
        
        row_idx = None
        for row_idx in range(self.data_start_row, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=self.barcode_column).value
            if cell_value and str(cell_value).strip() == barcode:
                break
        else:
            wb.close()
            return False
        
        # 업데이트할 컬럼들
        for key, value in updates.items():
            if key in self.extra_columns:
                col_idx = self.extra_columns[key]
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 저장
        wb.save(file_path)
        wb.close()
        return True
    
    def read_all_items(self):
        """모든 항목 읽기"""
        import openpyxl
        
        file_path = self.get_file_path()
        if not file_path.exists():
            return []
        
        # data_only=True: 수식 대신 계산된 값을 읽음
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[self.sheet_name]
        
        items = []
        current_item = None
        
        for row_idx in range(self.data_start_row, ws.max_row + 1):
            barcode = ws.cell(row=row_idx, column=self.barcode_column).value
            name = ws.cell(row=row_idx, column=self.name_column).value
            
            # 바코드가 있으면 새 항목
            if barcode:
                if current_item:
                    items.append(current_item)
                
                current_item = {
                    'row': row_idx,
                    'barcode': str(barcode).strip(),
                    'name': str(name).strip() if name else ''
                }
                
                # 추가 컬럼 읽기 (수식이 계산된 값으로 읽힘)
                for key, col_idx in self.extra_columns.items():
                    value = ws.cell(row=row_idx, column=col_idx).value
                    # 숫자가 아닌 값은 0으로 처리
                    if value is not None and isinstance(value, (int, float)):
                        current_item[key] = value
                    elif value is not None and isinstance(value, str):
                        # 문자열이면 숫자로 변환 시도
                        try:
                            current_item[key] = float(value)
                        except ValueError:
                            current_item[key] = 0
                    else:
                        current_item[key] = 0
            
            # 이름만 있으면 이전 항목에 추가 (여러 행에 걸친 이름)
            elif current_item and name:
                current_item['name'] += ' ' + str(name).strip()
        
        if current_item:
            items.append(current_item)
        
        wb.close()
        
        # 총 항목 수 업데이트
        self.total_items = len(items)
        self.save(update_fields=['total_items'])
        
        return items


class ExcelUpdateLog(models.Model):
    """엑셀 파일 업데이트 로그"""
    
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)
    document = models.ForeignKey(
        ExcelMasterDocument,
        on_delete=models.CASCADE,
        related_name='update_logs',
        verbose_name=_('문서')
    )
    barcode = models.CharField(_('바코드'), max_length=100)
    action = models.CharField(_('작업'), max_length=50)  # scan, update, add
    updates = models.JSONField(_('변경내용'), default=dict)
    previous_values = models.JSONField(_('이전값'), default=dict, blank=True)
    
    created_at = models.DateTimeField(_('작업일시'), auto_now_add=True)
    created_by = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        verbose_name=_('작업자')
    )
    
    class Meta:
        db_table = 'excel_update_logs'
        verbose_name = _('엑셀 업데이트 로그')
        verbose_name_plural = _('엑셀 업데이트 로그')
        ordering = ['-created_at']
    
    def __str__(self):
        return f"{self.barcode} - {self.action} - {self.created_at}"
