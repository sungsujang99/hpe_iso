"""
엑셀 파일 생성 서비스
템플릿 기반으로 데이터를 채워서 엑셀 파일 생성
"""
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
from django.conf import settings
from django.core.files import File
import tempfile


class ExcelGenerator:
    """템플릿 기반 엑셀 파일 생성기"""
    
    def __init__(self, template):
        """
        Args:
            template: DocumentTemplate 인스턴스 (template_file이 있어야 함)
        """
        self.template = template
    
    def generate_from_data(self, data):
        """
        템플릿 파일을 복사하고 데이터를 채워서 새 엑셀 파일 생성
        
        Args:
            data: dict - content_data 형식의 데이터
            
        Returns:
            File 객체 (Django File)
        """
        if not self.template.template_file:
            raise ValueError("템플릿 파일이 없습니다.")
        
        # 템플릿 파일 경로
        template_path = self.template.template_file.path
        
        # 워크북 열기
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 데이터 매핑 규칙 (템플릿에 따라 다를 수 있음)
        # 기본적으로 fields_schema에 정의된 필드를 찾아서 채움
        if hasattr(self.template, 'fields_schema') and self.template.fields_schema:
            self._fill_template_with_schema(ws, data)
        else:
            # 스키마가 없으면 기본 매핑
            self._fill_template_simple(ws, data)
        
        # 임시 파일에 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        wb.close()
        
        # Django File 객체로 반환
        with open(temp_file.name, 'rb') as f:
            django_file = File(f)
            content = django_file.read()
        
        # 임시 파일 삭제
        Path(temp_file.name).unlink()
        
        # 메모리에서 File 객체 생성
        from io import BytesIO
        file_obj = BytesIO(content)
        return File(file_obj, name=f'{self.template.name}.xlsx')
    
    def _fill_template_with_schema(self, ws, data):
        """
        fields_schema에 정의된 필드 기반으로 데이터 채우기
        
        스키마 예시:
        {
            'audit_date': {'type': 'date', 'label': '심사일자', 'cell': 'B5'},
            'auditor': {'type': 'text', 'label': '심사원', 'cell': 'B6'},
        }
        """
        fields_schema = self.template.fields_schema
        
        # 각 필드에 대해
        for field_name, field_config in fields_schema.items():
            if field_name in data:
                value = data[field_name]
                
                # 셀 위치가 지정된 경우
                if 'cell' in field_config:
                    cell_ref = field_config['cell']
                    try:
                        # 병합된 셀인지 확인
                        cell = ws[cell_ref]
                        if hasattr(cell, 'value'):
                            cell.value = value
                        else:
                            # MergedCell인 경우 병합 해제 후 값 입력
                            print(f"⚠️  {cell_ref}는 병합된 셀입니다. 건너뜁니다.")
                    except Exception as e:
                        print(f"⚠️  셀 {cell_ref}에 값을 입력할 수 없습니다: {e}")
                else:
                    # 셀 위치가 없으면 라벨을 찾아서 그 옆에 채움
                    self._find_and_fill_by_label(ws, field_config.get('label'), value)
    
    def _fill_template_simple(self, ws, data):
        """
        간단한 매핑: 키-값 쌍을 순서대로 채우기
        첫 번째 열에 키, 두 번째 열에 값
        """
        start_row = 3  # 헤더 건너뛰기
        
        for i, (key, value) in enumerate(data.items(), start=start_row):
            ws.cell(row=i, column=1).value = key
            ws.cell(row=i, column=2).value = value
    
    def _find_and_fill_by_label(self, ws, label, value):
        """
        워크시트에서 라벨을 찾아서 그 옆 셀에 값을 채우기
        """
        if not label:
            return
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip() == label.strip():
                    # 라벨 오른쪽 셀에 값 채우기
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    target_cell.value = value
                    return
