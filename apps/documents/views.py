"""
Document Views
"""
from rest_framework import viewsets, status, generics, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import models
from django.utils import timezone
from django.http import FileResponse
from django.shortcuts import get_object_or_404

from apps.accounts.permissions import IsAdminRole, IsManagerOrAdmin, CanReviewDocument, CanApproveDocument
from .models import (
    DocumentCategory, DocumentTemplate, Document,
    DocumentComment, DocumentHistory, DocumentAttachment
)
from .serializers import (
    DocumentCategorySerializer, DocumentTemplateSerializer,
    DocumentListSerializer, DocumentDetailSerializer,
    DocumentCreateSerializer, DocumentUpdateSerializer,
    DocumentSubmitSerializer, DocumentReviewSerializer,
    DocumentApprovalSerializer, DocumentCommentSerializer,
    DocumentAttachmentSerializer
)
from .services import PDFGenerator


def get_client_ip(request):
    """클라이언트 IP 주소 추출"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_document_history(document, user, action, from_status='', to_status='', comment='', ip_address=None):
    """문서 이력 기록"""
    DocumentHistory.objects.create(
        document=document,
        user=user,
        action=action,
        from_status=from_status,
        to_status=to_status,
        comment=comment,
        ip_address=ip_address
    )


class DocumentCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    """문서 카테고리 ViewSet"""
    
    queryset = DocumentCategory.objects.filter(is_active=True)
    serializer_class = DocumentCategorySerializer
    permission_classes = [IsAuthenticated]


class DocumentTemplateViewSet(viewsets.ReadOnlyModelViewSet):
    """문서 템플릿 ViewSet"""
    
    queryset = DocumentTemplate.objects.filter(is_active=True)
    serializer_class = DocumentTemplateSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        category_id = self.request.query_params.get('category')
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        return queryset


class DocumentViewSet(viewsets.ModelViewSet):
    """문서 관리 ViewSet"""
    
    queryset = Document.objects.select_related(
        'category', 'template', 'created_by', 'reviewed_by', 'approved_by'
    )
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return DocumentCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return DocumentUpdateSerializer
        elif self.action == 'retrieve':
            return DocumentDetailSerializer
        return DocumentListSerializer
    
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # 필터링
        status_filter = self.request.query_params.get('status')
        category = self.request.query_params.get('category')
        search = self.request.query_params.get('search')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if category:
            queryset = queryset.filter(category_id=category)
        if search:
            queryset = queryset.filter(
                models.Q(document_number__icontains=search) |
                models.Q(title__icontains=search)
            )
        
        # 역할별 접근 제어
        if user.is_admin:
            return queryset
        elif user.is_manager:
            # 검토자: 본인 작성 + 검토 대기 문서
            return queryset.filter(
                models.Q(created_by=user) |
                models.Q(status='pending_review')
            )
        else:
            # 일반 사용자: 본인 작성 문서만
            return queryset.filter(created_by=user)
    
    def perform_create(self, serializer):
        document = serializer.save()
        
        # 엑셀 템플릿이 있는 경우: 셀 편집 데이터로 엑셀 파일 생성
        if document.template and document.template.template_file and not document.excel_file:
            template_file_name = str(document.template.template_file)
            is_excel_template = template_file_name.endswith('.xlsx') or template_file_name.endswith('.xls')
            
            if is_excel_template and document.content_data and '_excel_cells' in document.content_data:
                try:
                    import openpyxl
                    import shutil
                    import tempfile
                    from pathlib import Path
                    from django.conf import settings
                    from django.core.files import File
                    
                    # 템플릿 파일 복사
                    template_path = Path(settings.MEDIA_ROOT) / str(document.template.template_file)
                    
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                        shutil.copy2(template_path, tmp.name)
                        
                        # openpyxl로 셀 수정
                        wb = openpyxl.load_workbook(tmp.name)
                        sheet_name = document.content_data.get('_sheet_name', wb.sheetnames[0])
                        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                        
                        # 사용자 편집 셀 반영
                        for cell_data in document.content_data['_excel_cells']:
                            row = cell_data['row'] + 1  # 0-indexed -> 1-indexed
                            col = cell_data['col'] + 1
                            value = cell_data['value']
                            
                            try:
                                cell = ws.cell(row=row, column=col)
                                # 숫자 변환 시도
                                try:
                                    value = int(value)
                                except (ValueError, TypeError):
                                    try:
                                        value = float(value)
                                    except (ValueError, TypeError):
                                        pass
                                cell.value = value
                            except Exception:
                                pass
                        
                        # 작성자 이름 자동 입력 (K3 셀 - 작성란)
                        try:
                            user = self.request.user
                            user_name = user.get_full_name() or user.username
                            ws['K3'] = user_name
                        except Exception:
                            pass
                        
                        wb.save(tmp.name)
                        
                        # 저장
                        filename = f'{document.document_number}.xlsx'
                        with open(tmp.name, 'rb') as f:
                            document.excel_file.save(filename, File(f), save=True)
                    
                    # content_data에서 임시 데이터 정리
                    clean_data = {k: v for k, v in document.content_data.items() if not k.startswith('_')}
                    document.content_data = clean_data
                    document.save(update_fields=['content_data'])
                    
                    print(f"✅ 엑셀 파일 생성 (셀 편집): {filename}")
                except Exception as e:
                    print(f"⚠️  엑셀 파일 생성 실패: {str(e)}")
            else:
                # 기존 방식: fields_schema 기반 생성
                try:
                    from .excel_generator import ExcelGenerator
                    
                    generator = ExcelGenerator(document.template)
                    excel_file = generator.generate_from_data(document.content_data)
                    
                    filename = f'{document.document_number}.xlsx'
                    document.excel_file.save(filename, excel_file, save=True)
                    
                    print(f"✅ 엑셀 파일 자동 생성: {filename}")
                except Exception as e:
                    print(f"⚠️  엑셀 파일 생성 실패: {str(e)}")
        
        log_document_history(
            document=document,
            user=self.request.user,
            action='문서 생성',
            to_status=document.status,
            ip_address=get_client_ip(self.request)
        )
    
    def perform_update(self, serializer):
        document = serializer.save()
        log_document_history(
            document=document,
            user=self.request.user,
            action='문서 수정',
            ip_address=get_client_ip(self.request)
        )
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """문서 제출 (승인 요청)"""
        document = self.get_object()
        
        if not document.can_submit:
            return Response(
                {'error': '제출할 수 없는 상태입니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = DocumentSubmitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        old_status = document.status
        
        # 부서장이 작성한 문서는 검토 단계를 건너뛰고 바로 승인 대기
        if request.user.is_department_head:
            document.status = Document.Status.PENDING_APPROVAL
            document.reviewed_by = request.user  # 자동 검토 완료 처리
            document.reviewed_at = timezone.now()
            message = '문서가 최종 승인 대기 상태로 제출되었습니다. (검토 자동 완료)'
            action_text = '승인 요청 (부서장, 검토 생략)'
        else:
            document.status = Document.Status.PENDING_REVIEW
            message = '문서가 검토 요청되었습니다.'
            action_text = '검토 요청'
        
        document.submitted_at = timezone.now()
        document.save(update_fields=['status', 'submitted_at', 'reviewed_by', 'reviewed_at'])
        
        # 코멘트 저장
        if serializer.validated_data.get('comment'):
            DocumentComment.objects.create(
                document=document,
                user=request.user,
                comment_type='general',
                content=serializer.validated_data['comment']
            )
        
        log_document_history(
            document=document,
            user=request.user,
            action=action_text,
            from_status=old_status,
            to_status=document.status,
            comment=serializer.validated_data.get('comment', ''),
            ip_address=get_client_ip(request)
        )
        
        return Response({
            'message': message,
            'status': document.status
        })
    
    @action(detail=True, methods=['post'], permission_classes=[CanReviewDocument])
    def review(self, request, pk=None):
        """문서 검토 (부서장 전용)"""
        document = self.get_object()
        
        if not document.can_review:
            return Response(
                {'error': '검토할 수 없는 상태입니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 관리자가 아닌 경우, 같은 부서의 부서장만 검토 가능
        if not request.user.is_admin:
            if not request.user.is_department_head:
                return Response(
                    {'error': '부서장만 검토할 수 있습니다.'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # 작성자와 같은 부서인지 확인
            if document.created_by.department != request.user.department:
                return Response(
                    {'error': '본인 부서의 문서만 검토할 수 있습니다.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        serializer = DocumentReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        old_status = document.status
        action_type = serializer.validated_data['action']
        comment = serializer.validated_data.get('comment', '')
        
        if action_type == 'approve':
            document.status = Document.Status.PENDING_APPROVAL
            document.reviewed_by = request.user
            document.reviewed_at = timezone.now()
            action_text = '검토 승인'
            comment_type = 'review'
        else:  # reject
            document.status = Document.Status.REJECTED
            document.reviewed_by = request.user
            document.reviewed_at = timezone.now()
            action_text = '검토 반려'
            comment_type = 'rejection'
        
        document.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])
        
        # 코멘트 저장
        if comment:
            DocumentComment.objects.create(
                document=document,
                user=request.user,
                comment_type=comment_type,
                content=comment
            )
        
        log_document_history(
            document=document,
            user=request.user,
            action=action_text,
            from_status=old_status,
            to_status=document.status,
            comment=comment,
            ip_address=get_client_ip(request)
        )
        
        return Response({
            'message': f'문서가 {action_text}되었습니다.',
            'status': document.status
        })
    
    @action(detail=True, methods=['post'], permission_classes=[CanApproveDocument])
    def approve(self, request, pk=None):
        """문서 최종 승인 (관리자 전용)"""
        document = self.get_object()
        
        if not document.can_approve:
            return Response(
                {'error': '승인할 수 없는 상태입니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = DocumentApprovalSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        old_status = document.status
        action_type = serializer.validated_data['action']
        comment = serializer.validated_data.get('comment', '')
        
        if action_type == 'approve':
            document.status = Document.Status.APPROVED
            document.approved_by = request.user
            document.approved_at = timezone.now()
            document.is_locked = True  # 승인 후 잠금
            action_text = '최종 승인'
            comment_type = 'approval'
            
            # PDF 생성
            try:
                pdf_generator = PDFGenerator()
                pdf_file = pdf_generator.generate_document_pdf(document)
                document.pdf_file = pdf_file
            except Exception as e:
                # PDF 생성 실패해도 승인은 진행
                pass
        else:  # reject
            document.status = Document.Status.REJECTED
            document.approved_by = request.user
            document.approved_at = timezone.now()
            action_text = '최종 반려'
            comment_type = 'rejection'
        
        document.save(update_fields=['status', 'approved_by', 'approved_at', 'is_locked', 'pdf_file'])
        
        # 코멘트 저장
        if comment:
            DocumentComment.objects.create(
                document=document,
                user=request.user,
                comment_type=comment_type,
                content=comment
            )
        
        log_document_history(
            document=document,
            user=request.user,
            action=action_text,
            from_status=old_status,
            to_status=document.status,
            comment=comment,
            ip_address=get_client_ip(request)
        )
        
        return Response({
            'message': f'문서가 {action_text}되었습니다.',
            'status': document.status
        })
    
    @action(detail=True, methods=['post'])
    def revise(self, request, pk=None):
        """반려된 문서 재작성"""
        document = self.get_object()
        
        if document.status != Document.Status.REJECTED:
            return Response(
                {'error': '반려된 문서만 재작성할 수 있습니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if document.created_by != request.user:
            return Response(
                {'error': '본인이 작성한 문서만 재작성할 수 있습니다.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        old_status = document.status
        document.status = Document.Status.DRAFT
        document.reviewed_by = None
        document.reviewed_at = None
        document.approved_by = None
        document.approved_at = None
        document.revision = str(int(document.revision) + 1)
        document.save()
        
        log_document_history(
            document=document,
            user=request.user,
            action='문서 재작성',
            from_status=old_status,
            to_status=document.status,
            ip_address=get_client_ip(request)
        )
        
        return Response({
            'message': '문서를 재작성할 수 있습니다.',
            'revision': document.revision
        })
    
    @action(detail=False, methods=['get'])
    def related_documents(self, request):
        """템플릿과 관련된 문서 목록 조회"""
        template_id = request.query_params.get('template')
        if not template_id:
            return Response({'error': '템플릿 ID가 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            template = DocumentTemplate.objects.get(id=template_id)
        except DocumentTemplate.DoesNotExist:
            return Response({'error': '템플릿을 찾을 수 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        
        # description에서 매뉴얼 섹션 추출
        section = None
        if '매뉴얼 섹션:' in template.description:
            section = template.description.split('매뉴얼 섹션:')[1].strip().split('\n')[0].strip()
        
        # 같은 섹션에 속한 관련 문서 찾기
        related_docs = []
        if section:
            related_templates = DocumentTemplate.objects.filter(
                description__contains=f'매뉴얼 섹션: {section}'
            ).exclude(id=template_id)
            
            for rel_template in related_templates:
                # 해당 템플릿으로 승인된 최신 문서 찾기
                doc = Document.objects.filter(
                    template=rel_template,
                    status=Document.Status.APPROVED
                ).order_by('-approved_at').first()
                
                if doc:
                    related_docs.append({
                        'id': str(doc.id),
                        'document_number': doc.document_number,
                        'title': doc.title,
                        'template_name': rel_template.name,
                        'content_data': doc.content_data,
                        'approved_at': doc.approved_at
                    })
        
        return Response({
            'template': template.name,
            'section': section,
            'related_documents': related_docs
        })
    
    @action(detail=False, methods=['post'])
    def extract_shared_data(self, request):
        """관련 문서들에서 공유 데이터 추출"""
        template_id = request.data.get('template')
        if not template_id:
            return Response({'error': '템플릿 ID가 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            template = DocumentTemplate.objects.get(id=template_id)
        except DocumentTemplate.DoesNotExist:
            return Response({'error': '템플릿을 찾을 수 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        
        # description에서 매뉴얼 섹션 추출
        section = None
        if '매뉴얼 섹션:' in template.description:
            section = template.description.split('매뉴얼 섹션:')[1].strip().split('\n')[0].strip()
        
        shared_data = {}
        
        # 같은 섹션의 승인된 문서들에서 데이터 수집
        if section:
            related_templates = DocumentTemplate.objects.filter(
                description__contains=f'매뉴얼 섹션: {section}'
            ).exclude(id=template_id)
            
            for rel_template in related_templates:
                doc = Document.objects.filter(
                    template=rel_template,
                    status=Document.Status.APPROVED
                ).order_by('-approved_at').first()
                
                if doc and doc.content_data:
                    # 목적, 적용범위 등 공통 필드 수집
                    for key in ['purpose', 'scope', 'responsibility', 'procedure']:
                        if key in doc.content_data and doc.content_data[key]:
                            if key not in shared_data:
                                shared_data[key] = []
                            shared_data[key].append({
                                'source': doc.document_number,
                                'content': doc.content_data[key]
                            })
        
        # 회사 정보 추가 (항상 포함)
        shared_data['company_info'] = {
            'company_name': '주식회사 에이치피엔지니어링',
            'company_name_en': 'HP Engineering Co., Ltd',
            'company_address': '경상남도 창원시 의창구 원이대로 271, 3층 322호(봉곡동, 한마음타워)',
        }
        
        # 품질 방침 추가 (매뉴얼에서)
        if template.category.code == 'HP-QM':
            shared_data['quality_policy'] = {
                'policy': '고객 요구사항 준수를 통한 고객만족 실현',
                'objectives': [
                    '고객과 소통하는 경영',
                    '실행중심·고객중심·인간중심',
                    '고객약속 최우선'
                ]
            }
        
        return Response({
            'template': template.name,
            'section': section,
            'shared_data': shared_data
        })
    
    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """PDF 다운로드"""
        document = self.get_object()
        
        if not document.pdf_file:
            # PDF가 없으면 생성
            try:
                pdf_generator = PDFGenerator()
                pdf_file = pdf_generator.generate_document_pdf(document)
                document.pdf_file = pdf_file
                document.save(update_fields=['pdf_file'])
            except Exception as e:
                return Response(
                    {'error': f'PDF 생성 중 오류가 발생했습니다: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        
        return FileResponse(
            document.pdf_file.open('rb'),
            as_attachment=True,
            filename=f'{document.document_number}.pdf'
        )
    
    @action(detail=True, methods=['get'])
    def preview_pdf(self, request, pk=None):
        """PDF 미리보기"""
        document = self.get_object()
        
        if not document.pdf_file:
            try:
                pdf_generator = PDFGenerator()
                pdf_file = pdf_generator.generate_document_pdf(document)
                document.pdf_file = pdf_file
                document.save(update_fields=['pdf_file'])
            except Exception as e:
                return Response(
                    {'error': f'PDF 생성 중 오류가 발생했습니다: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        
        return FileResponse(
            document.pdf_file.open('rb'),
            content_type='application/pdf'
        )


class DocumentAttachmentViewSet(viewsets.ModelViewSet):
    """첨부파일 관리 ViewSet"""
    
    serializer_class = DocumentAttachmentSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        document_id = self.kwargs.get('document_pk')
        return DocumentAttachment.objects.filter(document_id=document_id)
    
    def perform_create(self, serializer):
        document_id = self.kwargs.get('document_pk')
        document = get_object_or_404(Document, pk=document_id)
        
        if document.is_locked:
            raise serializers.ValidationError('잠긴 문서에는 파일을 첨부할 수 없습니다.')
        
        serializer.save(
            document=document,
            uploaded_by=self.request.user
        )


class PendingReviewListView(generics.ListAPIView):
    """검토 대기 문서 목록"""
    
    serializer_class = DocumentListSerializer
    permission_classes = [IsAuthenticated]  # 모든 로그인 사용자 접근 가능
    
    def get_queryset(self):
        user = self.request.user
        queryset = Document.objects.filter(
            status=Document.Status.PENDING_REVIEW
        ).select_related('category', 'created_by')
        
        # 역할별 필터링
        if user.is_admin:
            return queryset  # 관리자는 모든 문서
        elif user.is_manager:
            return queryset  # 중간관리자는 검토 가능한 모든 문서
        else:
            return queryset.filter(created_by=user)  # 일반 사용자는 본인이 작성한 문서만


class PendingApprovalListView(generics.ListAPIView):
    """승인 대기 문서 목록"""
    
    serializer_class = DocumentListSerializer
    permission_classes = [IsAuthenticated]  # 모든 로그인 사용자 접근 가능
    
    def get_queryset(self):
        user = self.request.user
        queryset = Document.objects.filter(
            status=Document.Status.PENDING_APPROVAL
        ).select_related('category', 'created_by', 'reviewed_by')
        
        # 역할별 필터링
        if user.is_admin:
            return queryset  # 관리자는 모든 문서
        elif user.is_manager:
            return queryset.filter(created_by=user)  # 중간관리자는 본인이 작성한 문서만
        else:
            return queryset.filter(created_by=user)  # 일반 사용자는 본인이 작성한 문서만
